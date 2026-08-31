"""
Import d'un programme annuel de production depuis un fichier Excel.
Lit la matrice postes (lignes) × mois (colonnes) et remplit
la table productions_mensuelles.
Validé sur le format Prog_ann_ex.xlsx (Budget Bouchane).
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from sqlalchemy.orm import Session

from app.models.production_mensuelle import ProductionMensuelle


def importer_production(fichier_bytes: bytes, programme_id: int, db: Session) -> dict:
    wb = openpyxl.load_workbook(BytesIO(fichier_bytes), data_only=True)
    ws = wb.active

    # Unité (cellule C5 contient "Unité :Ktsm")
    unite_cell = ws.cell(row=5, column=3).value
    unite = "Ktsm"
    if unite_cell and ":" in str(unite_cell):
        unite = str(unite_cell).split(":")[-1].strip()

    # Colonnes des mois : ligne 5, colonnes 5 à 16
    mois_cols = {}
    for col in range(5, 17):
        v = ws.cell(row=5, column=col).value
        if isinstance(v, datetime):
            mois_cols[col] = v.date()

    if not mois_cols:
        return {"importes": 0, "erreur": "Aucune colonne de mois détectée (ligne 5)."}

    # Supprime l'ancien import pour ce programme (ré-import propre)
    db.query(ProductionMensuelle).filter(
        ProductionMensuelle.programme_id == programme_id
    ).delete()

    section_courante = None
    nb_lignes = 0
    nb_valeurs = 0

    for row in range(9, 46):
        # Section en colonne B
        section_b = ws.cell(row=row, column=2).value
        if section_b and str(section_b).strip():
            section_courante = str(section_b).strip()

        # Poste en colonne C
        poste = ws.cell(row=row, column=3).value
        if not poste or not str(poste).strip():
            continue
        poste = str(poste).strip()

        ligne_a_des_valeurs = False
        for col, mois in mois_cols.items():
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                prod = ProductionMensuelle(
                    programme_id=programme_id,
                    section=section_courante,
                    poste=poste,
                    mois=mois,
                    unite=unite,
                    valeur_prevue=round(float(val), 2),
                )
                db.add(prod)
                nb_valeurs += 1
                ligne_a_des_valeurs = True

        if ligne_a_des_valeurs:
            nb_lignes += 1

    db.commit()
    return {
        "importes": nb_lignes,
        "valeurs": nb_valeurs,
        "mois": len(mois_cols),
        "unite": unite,
    }