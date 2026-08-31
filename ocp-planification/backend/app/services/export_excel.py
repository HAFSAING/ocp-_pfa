from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.panneau import Panneau
from app.models.tranchee import Tranchee
from app.models.saisie_mensuelle import SaisieMensuelle
from app.models.resultat import Resultat
from app.models.programme import Programme


def exporter_programme(programme_id: int, db: Session) -> BytesIO:
    programme = db.query(Programme).filter(Programme.id == programme_id).first()

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    header_fill = PatternFill(start_color="163E2C", end_color="163E2C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(color="163E2C", bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:Q1")
    cell = ws["A1"]
    cell.value = f"OCP — Planning {programme.mine or ''} — Année {programme.annee}"
    cell.font = title_font
    ws.row_dimensions[1].height = 24

    headers = [
        "Ordre", "Panneau", "Tranchée", "Profil", "État", "Long (m)", "Larg (m)", "H (m)",
        "P.stérile", "P.phosphate", "Surface (m²)", "Vol. stérile", "Vol. phosphate",
        "Tonnage TSM", "HMB (h)", "ML à forer", "Jours prévus",
    ]
    row_start = 3
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row_start, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    ETAT_LABELS = {"non_commence": "Non commencé", "en_cours": "En cours", "epuise": "Épuisé"}

    tranchees = (
        db.query(Tranchee)
        .join(Panneau, Tranchee.panneau_id == Panneau.id)
        .filter(Panneau.programme_id == programme_id)
        .order_by(Tranchee.ordre_execution.asc().nullslast())
        .all()
    )

    row = row_start + 1
    for t in tranchees:
        resultat = (
            db.query(Resultat)
            .join(SaisieMensuelle, Resultat.saisie_id == SaisieMensuelle.id)
            .filter(SaisieMensuelle.tranchee_id == t.id)
            .order_by(Resultat.date_calcul.desc())
            .first()
        )
        r = resultat

        valeurs = [
            t.ordre_execution or "",
            t.panneau.code_pan if t.panneau else "",
            t.code,
            t.profil or "",
            ETAT_LABELS.get(t.etat.value, t.etat.value),
            t.longueur_m or "",
            t.largeur_m or "",
            t.hauteur_m or "",
            t.puissance_sterile_m or "",
            t.puissance_phosphate_m or "",
            r.surface if r else "",
            r.volume_sterile if r else "",
            r.volume_phosphate if r else "",
            r.tonnage_phosphate_tsm if r else "",
            r.hmb if r else "",
            r.ml_a_forer if r else "",
            r.jours_prevus if r else "",
        ]
        for col, v in enumerate(valeurs, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if col >= 6:
                c.alignment = center
        row += 1

    largeurs = [7, 14, 12, 8, 13, 9, 9, 8, 11, 12, 12, 13, 14, 12, 10, 11, 12]
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer